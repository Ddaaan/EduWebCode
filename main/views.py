from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.http import JsonResponse
from django.views.generic import View

from django.contrib.auth import login, authenticate
from django.contrib import messages
from django.contrib.auth.forms import AuthenticationForm

from django.shortcuts import render, redirect, get_object_or_404
from main.models import School, File, FileView
from .models import Post, PostView
from django.utils import timezone
from .forms import LoginForm

import pandas as pd
import openpyxl
import os

import json
from django.http import JsonResponse
from django.template.loader import render_to_string

from django.http import FileResponse, Http404
import mimetypes
from django.conf import settings

from django.http import HttpResponseForbidden, HttpResponseNotAllowed

from django.contrib.auth.models import User  # Django User 모델
from django.contrib.auth.hashers import check_password

from django.db.models.functions import Lower

from django.contrib.auth.decorators import login_required

import queue
import threading

response_queue = queue.Queue()

# Create your views here.
def main_index(request):
    return render(request, "mainpage.html")

def post(request):
    return render(request, 'post.html')

def about1(request):
    return render(request, 'about1.html')

def about2(request):
    return render(request, 'about2.html')

def file(request):
    return render(request, 'file.html')

def survey_complete(request):
    return render(request, 'survey_complete.html')

@login_required(login_url='/jns/admin-login/')
def statistics_admin_page(request):
    # role 값을 세션에서 가져옴
    role = request.session.get('role')
    
    # 모든 지역을 가져오되, 지역청/학교 관리자는 본인의 지역만 보게 설정
    if request.session.get('role') in ['regional_admin', 'school_admin']:
        # 지역청 관리자의 지역을 세션에서 가져옴
        region = request.session.get('region')
        regions = [region]  # 지역청/학교 관리자는 본인의 지역만 선택 가능
    else:
        # 본청 관리자는 모든 지역을 볼 수 있음
        regions = School.objects.values_list('district', flat=True).distinct().order_by('district')
    
    # GET 요청일 때 지역 데이터만 템플릿으로 전달
    return render(request, 'statistics_admin_page.html', {'regions': regions, 'role': role})

@login_required(login_url='/jns/admin-login/')
def statistics_admin_region_page(request):
    # role 값을 세션에서 가져옴
    role = request.session.get('role')
    
    # 모든 지역을 가져오되, 지역청/학교 관리자는 본인의 지역만 보게 설정
    if request.session.get('role') in ['regional_admin', 'school_admin']:
        # 지역청 관리자의 지역을 세션에서 가져옴
        region = request.session.get('region')
        regions = [region]  # 지역청/학교 관리자는 본인의 지역만 선택 가능
    else:
        # 본청 관리자는 모든 지역을 볼 수 있음
        regions = School.objects.values_list('district', flat=True).distinct().order_by('district')
    
    # GET 요청일 때 지역 데이터만 템플릿으로 전달
    return render(request, 'statistics_admin_region_page.html', {'regions': regions, 'role': role})

@login_required(login_url='/jns/admin-login/')
def statistics_admin_total_page(request):
    # role 값을 세션에서 가져옴
    role = request.session.get('role')
    
    # 모든 지역을 가져오되, 지역청/학교 관리자는 본인의 지역만 보게 설정
    if request.session.get('role') in ['regional_admin', 'school_admin']:
        # 지역청 관리자의 지역을 세션에서 가져옴
        region = request.session.get('region')
        regions = [region]  # 지역청/학교 관리자는 본인의 지역만 선택 가능
    else:
        # 본청 관리자는 모든 지역을 볼 수 있음
        regions = School.objects.values_list('district', flat=True).distinct().order_by('district')
    
    # GET 요청일 때 지역 데이터만 템플릿으로 전달
    return render(request, 'statistics_admin_total_page.html', {'regions': regions, 'role': role})


from django.shortcuts import render, redirect
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password, check_password
from .forms import PasswordChangeForm
from .models import School

from django.contrib.auth import authenticate, login, logout

@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.POST)
        if form.is_valid():
            old_password = form.cleaned_data.get('old_password')
            new_password = form.cleaned_data.get('new_password')
            confirm_password = form.cleaned_data.get('confirm_password')

            # 현재 로그인된 학교의 ID를 세션에서 가져옴
            school_id = request.session.get('school_id')

            try:
                school = School.objects.get(school_id=school_id)
                if check_password(old_password, school.school_pw):
                    if new_password == confirm_password:
                        # 새로운 비밀번호 해쉬화 후 업데이트
                        school.school_pw = make_password(new_password)
                        school.save()

                        # Django User 모델 비밀번호 갱신
                        user = request.user
                        user.set_password(new_password)
                        user.save()

                        # 사용자 로그아웃
                        logout(request)
                        
                        # 로그아웃 후 로그인 페이지로 리다이렉트
                        messages.success(request, '비밀번호가 성공적으로 변경되었습니다. 다시 로그인 해주세요.')
                        return redirect('admin_login')
                    
                    else:
                        form.add_error('confirm_password', '새 비밀번호가 일치하지 않습니다.')
                else:
                    form.add_error('old_password', '현재 비밀번호가 일치하지 않습니다.')
            except School.DoesNotExist:
                form.add_error(None, '사용자를 찾을 수 없습니다.')
    else:
        form = PasswordChangeForm()

    return render(request, 'change_password.html', {'form': form})




# 관리자 로그인
def admin_login(request):
    # 이미 로그인이 되어있는지 세션을 확인
    if request.session.get('school_name') and request.session.get('role'):
        return redirect('admin_dashboard')  # 대시보드 페이지로 이동
    
    form = LoginForm(request.POST or None)
    
    if request.method == 'POST':
        if form.is_valid():
            school_id = form.cleaned_data.get('school_id')
            school_pw = form.cleaned_data.get('school_pw')
            
            try:
                #아이디와 비번 확인
                school = School.objects.get(school_id = school_id)
                
                #비번 일치 확인
                if check_password(school_pw, school.school_pw):  # 비번 암호화 되어있는 것 비교
                    request.session['school_id'] = school.school_id
                    request.session['school_name'] = school.school_name  # 이 부분에서 school_name이 저장됨
                    request.session['region'] = school.district
                    
                    # school_level에 따라 역할 설정
                    if school.school_level in ['초등학교', '중학교', '고등학교', '유치원', '특수학교', '각종학교(중)', '각종학교(고)']:
                        request.session['role'] = 'school_admin'
                    elif school.school_level == '지역청':
                        request.session['role'] = 'regional_admin'
                    elif school.school_level == '본청':
                        request.session['role'] = 'main_admin'
                        
                    # 세션 설정 후 로그인 처리 (Django User 모델을 사용해서 로그인)
                    user = authenticate(request, username=school.school_id, password=school_pw)
                    if user:
                        login(request, user)  # Django의 인증 시스템 사용
                    else:
                        # 유저가 없을 경우 Django User 객체를 만들고 로그인 처리
                        # 수정: 기존 유저가 있는지 확인하고 없으면 생성하도록 수정
                        user, created = User.objects.get_or_create(username=school.school_id)
                        if created:
                            user.set_password(school_pw)  # 비밀번호 설정
                            user.save()
                        # 비밀번호 설정 후 인증 및 로그인
                        user = authenticate(request, username=school.school_id, password=school_pw)
                        if user:
                            login(request, user)
                    
                    return redirect ('admin_dashboard') # --> 로그인 완료 후 이동할 페이지
                
                else:
                    messages.error(request, '비밀번호가 일치하지 않습니다.')
                    
            except School.DoesNotExist:
                messages.error(request, '존재하지 않는 아이디입니다.')
        
    return render(request, 'admin_login.html', {'form':form})


# 관리자 대시보드 (role에 따라)
def admin_dashboard(request):
    role = request.session.get('role')
    school_id = request.session.get('school_id')
    school_name = request.session.get('school_name')
    region = request.session.get('region')
    
    if not role:
        messages.error(request, '권한이 없습니다. 다시 로그인 해주세요.')
        return redirect('admin_login')
    
    if role == 'school_admin':
        # 학교 관리자 대시보드로 이동
        return render(request, 'school_dashboard.html', {'school_name': school_name, 'school_id':school_id, 'region':region})
    elif role == 'regional_admin':
        # 지역청 관리자 대시보드로 이동
        return render(request, 'regional_dashboard.html', {'school_name': school_name, 'school_id':school_id, 'region':region})
    elif role == 'main_admin':
        # 본청 관리자 대시보드로 이동
        return render(request, 'main_dashboard.html', {'school_name': school_name, 'school_id':school_id, 'region':region})
    
    # 만약 역할이 잘못되었다면 다시 로그인 페이지로
    messages.error(request, '권한이 없습니다. 다시 로그인 해주세요.')
    return redirect('admin_login')

import zipfile
from io import BytesIO

def download_survey_data(request):
    # 다운로드할 파일 목록
    file_paths = [
        "/eduWeb/surveySite/main/surveydata/survey_result_student_each.xlsx",
        "/eduWeb/surveySite/main/surveydata/survey_result_parents_each.xlsx",
        "/eduWeb/surveySite/main/surveydata/survey_result_teacher_each.xlsx"
    ]

    # 메모리 내에서 zip 파일을 생성하기 위해 BytesIO 사용
    zip_buffer = BytesIO()

    # zip 파일 생성
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        for file_path in file_paths:
            file_name = os.path.basename(file_path)  # 파일명 추출
            zip_file.write(file_path, file_name)  # ZIP 파일에 추가

    # 응답을 위한 ZIP 파일을 메모리로부터 가져오기
    zip_buffer.seek(0)

    # HTTP 응답으로 ZIP 파일 전송
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="survey_data.zip"'

    return response

# 로그아웃
def admin_logout(request):
    request.session.flush()  # 세션 데이터를 모두 삭제
    messages.success(request, '로그아웃 되었습니다.')
    return redirect('admin_login')


# 정보 선택
def info_page(request):
    role = request.GET.get('role', 'default')
    regions = School.objects.values_list('district', flat=True).distinct().order_by('district')

    role_messages = {
        'student' : '학생 정보를 선택하세요',
        'parent' : '학부모 정보를 선택하세요',
        'teacher' : '교원 정보를 선택하세요'
    }
    message = role_messages.get(role, '본인 정보를 선택하세요')
    
    if request.method == "POST":
        region = request.POST.get('region')
        school_level = request.POST.get('school-level')
        school_name = request.POST.get('school-name')
        school_id = request.POST.get('school-id')
        
        # 특수 학교 학생 설문조사 X
        if role == 'student' and school_level in ['특수학교', '각종학교(중)', '각종학교(고)']:
            messages.error(request, "진행중인 설문조사가 없습니다")
            return render(request, 'infopage.html', {'regions': regions, 'message': message})
        
        try:
            school = School.objects.get(district = region, school_level=school_level, school_name=school_name, school_id=school_id)
            
            if school is None:
                messages.error(request, "입력한 정보와 일치하는 학교가 없습니다.")
                return render(request, 'infopage.html', {'regions': regions, 'message': message})

            # 학교 정보가 일치하면 role에 따라 페이지 리디렉션
            # school_id를 세션에 저장
            request.session['school_id'] = school_id
            
            if role == 'student':
                if school_level in ['초등학교']:
                    return redirect('ele-student-s')
                elif school_level in ['중학교', '고등학교']:
                    return redirect('midhigh-student-s')
            elif role == 'parent':
                if school_level == '유치원':
                    return redirect('kinder-parents-s')
                elif school_level in ['초등학교', '중학교', '고등학교', '각종학교(중)', '각종학교(고)', '특수학교']:
                    return redirect('school-parents-s')
            elif role == 'teacher':
                if school_level == '유치원':
                    return redirect('kinder-teacher-s')
                elif school_level in ['초등학교', '중학교', '고등학교', '각종학교(중)', '각종학교(고)', '특수학교']:
                    return redirect('school-teacher-s')
            else:
                messages.error(request, "유효하지 않은 접근입니다.")
        except School.DoesNotExist:
            messages.error(request, "입력한 정보와 일치하는 학교가 없습니다.")
            return render(request, 'infopage.html', {'regions': regions, 'message': message})

    return render(request, 'infopage.html', {'regions': regions, 'message': message})

# 학교명 정렬
def get_school_names(request):
    region = request.GET.get('region')
    school_level = request.GET.get('school_level')
    
    schools = School.objects.filter(district=region, school_level=school_level).values('school_name')
    school_list = list(schools)
        
    return JsonResponse({'schools' : school_list})

# 학교 id 가져오는 함수
def get_school_id(request):
    region = request.GET.get('region')
    school_level = request.GET.get('school_level')
    school_name = request.GET.get('school_name')
    
    schoolId = School.objects.filter(district=region, school_level=school_level, school_name=school_name).values('school_id')
    school_id = list(schoolId)
    
    return JsonResponse({'schoolId' : school_id})

##공지사항
# IP 주소 가져오는 함수
def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

#게시글 목록
def post_list(request):
    posts = Post.objects.all().order_by('-created_at') #작성일 내림차순 정렬
    return render(request, 'post_list.html', {'posts' : posts})

#게시글 작성
def post_create(request):
    # main_admin만 접근 가능
    if not request.user.is_authenticated or request.user.username != 'A11420':
        return HttpResponseForbidden("이 페이지에 접근할 수 있는 권한이 없습니다.")
    
    if request.method == 'POST':
        title = request.POST['title']
        content = request.POST['content']
        file = request.FILES.get('file') #파일 가져오기
        
        # 디버깅: 파일이 제대로 처리되는지 확인
        print("파일 처리:", file)
        
        Post.objects.create(title=title, content=content, file=file, created_at=timezone.now()) #게시글 저장
        
        return redirect('post_list') #작성 후 목록으로 이동
    
    return render(request, 'post_create.html')

#게시글 보기
def post_detail(request, post_id):
    post = get_object_or_404(Post, id=post_id)
    
    client_ip = get_client_ip(request)
    
    if not PostView.objects.filter(post=post, ip_address=client_ip).exists(): #접속한 ip로 이미 조회했는지 확인
        post.views += 1 #조회수
        post.save()
        
        PostView.objects.create(post=post, ip_address=client_ip)
        
    return render(request, 'post_detail.html', {'post' : post})

#파일 다운로드 함수
def download_file(request, file_path):
    file_full_path = os.path.join(settings.MEDIA_ROOT, file_path)
    
    # 파일이 존재하지 않을 경우 404 오류 반환
    if not os.path.exists(file_full_path):
        raise Http404("File does not exist")
    
    # 파일 다운로드 응답
    response = FileResponse(open(file_full_path, 'rb'))
    response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_full_path)}"'
    return response

#게시글 삭제 함수
def post_delete(request, post_id):
    post = get_object_or_404(Post, id=post_id)

    # main_admin 또는 특정 사용자만 삭제 가능
    if not request.user.is_authenticated or request.user.username != 'A11420':
        return HttpResponseForbidden("이 페이지에 접근할 수 있는 권한이 없습니다.")

    # POST 요청일 때만 삭제 진행
    if request.method == 'POST':
        post.delete()
        return redirect('post_list')

    # GET 요청일 경우 처리
    return HttpResponseNotAllowed(['POST'], "이 작업은 POST 요청만 허용됩니다.")



###자료실 함수
# 파일 목록
def file_list(request):
    files = File.objects.all().order_by('-created_at')  # 작성일 내림차순 정렬
    return render(request, 'file_list.html', {'files': files})

# 파일 업로드
def file_create(request):
    if not request.user.is_authenticated or request.user.username != 'A11420':
        return HttpResponseForbidden("이 페이지에 접근할 수 있는 권한이 없습니다.")
    
    if request.method == 'POST':
        title = request.POST['title']
        content = request.POST['content']
        file = request.FILES.get('file')  # 파일 가져오기
        
        File.objects.create(title=title, content=content, file=file, created_at=timezone.now())
        return redirect('file_list')
    
    return render(request, 'file_create.html')

# 파일 상세 보기
def file_detail(request, file_id):
    file = get_object_or_404(File, id=file_id)
    
    client_ip = get_client_ip(request)
    
    if not FileView.objects.filter(file=file, ip_address=client_ip).exists(): #접속한 ip로 이미 조회했는지 확인
        file.views += 1 #조회수
        file.save()
        
        FileView.objects.create(file=file, ip_address=client_ip)
    
    return render(request, 'file_detail.html', {'file': file})

# 파일 삭제
def file_delete(request, file_id):
    file = get_object_or_404(File, id=file_id)

    if not request.user.is_authenticated or request.user.username != 'A11420':
        return HttpResponseForbidden("이 페이지에 접근할 수 있는 권한이 없습니다.")
    
    if request.method == 'POST':
        file.delete()
        return redirect('file_list')

    return HttpResponseNotAllowed(['POST'], "이 작업은 POST 요청만 허용됩니다.")

import datetime

file_lock = threading.Lock()

#큐 저장 함수
def process_queue():
    while True:
        role, responses, school_id, ip_address = response_queue.get()  # 큐에서 데이터 꺼내기
        print(f"role: { role}, id: {school_id}, ip: {ip_address} 꺼내기")
        
        if role == 'student':
            excel_file_path = "\\eduWeb\\surveySite\\main\\surveydata\\survey_result_student.xlsx"
            excel_file_path2 = "\\eduWeb\\surveySite\\main\\surveydata\\backup\\survey_result_student_each_backup.xlsx"
            question_count = 22
            people_count_col = 27
    
        elif role == 'parents':
            excel_file_path = "\\eduWeb\\surveySite\\main\\surveydata\\backup\\survey_result_parents.xlsx"
            excel_file_path2 = "\\eduWeb\\surveySite\\main\\surveydata\\backup\\survey_result_parents_each_backup.xlsx"
            question_count = 23
            people_count_col = 28
            
        elif role == 'teacher':
            excel_file_path = "\\eduWeb\\surveySite\\main\\surveydata\\backup\\survey_result_teacher.xlsx"
            excel_file_path2 = "\\eduWeb\\surveySite\\main\\surveydata\\backup\\survey_result_teacher_each_backup.xlsx"
            question_count = 25
            people_count_col = 30
        
        # 응답을 텍스트 파일에 저장
        try:
            txt_file_path = "\\eduWeb\\surveySite\\main\\surveydata\\responses_log.txt"
            with open(txt_file_path, "a", encoding="utf-8") as txt_file:
                timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                txt_file.write(f"{role}    {timestamp}    {school_id}    {ip_address}   {responses}\n")
        except Exception as e:
            print(f"텍스트 파일에 저장 중 오류 발생: {e}")
        
        with file_lock:
            # 엑셀 파일 열기 및 응답 저장
            try:
                print("엑셀 파일 열기 시도...")
                wb = openpyxl.load_workbook(excel_file_path, data_only=True)
                ws = wb.active
                print("엑셀 파일 활성화")

                # 학교별 응답 저장
                row_to_update = None
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
                    if str(row[0].value).strip() == str(school_id).strip():
                        row_to_update = row[0].row
                        break

                if row_to_update:
                    for i, response in enumerate(responses, start=5):
                        current_value = ws.cell(row=row_to_update, column=i).value or 0
                        ws.cell(row=row_to_update, column=i).value = current_value + response

                    current_value = ws.cell(row=row_to_update, column=people_count_col).value or 0
                    ws.cell(row=row_to_update, column=people_count_col).value = current_value + 1
                    wb.save(excel_file_path)
                
                else:
                    print(f"학교 ID {school_id}에 해당하는 행을 찾을 수 없습니다.")
                
                wb.close()
                
                print(f"학교 ID {school_id}에 저장 완료")
                
            except Exception as e:
                print(f"엑셀 처리 중 오류 발생 (학교별 응답 저장): {e}")
            
            try:
                # 개개인별 응답 저장
                print("개개인별 응답 저장 시작...")
                each_wb = openpyxl.load_workbook(excel_file_path2, data_only=True)
                each_ws = each_wb.active
                print("개개인별 엑셀 파일 활성화")
                
                last_row = each_ws.max_row+1
                timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # 순번 매기기
                try:
                    num = int(each_ws.cell(row=last_row - 1, column=1).value) or 0
                except (TypeError, ValueError):
                    num = 0  # 숫자로 변환할 수 없는 경우 기본값 0
                    
                #print(f"no : {num + 1}, timestamp : {timestamp}, ip : {ip_address}, id : {school_id}") #디버깅

                each_ws.cell(row=last_row, column=1).value = num + 1
                each_ws.cell(row=last_row, column=2).value = timestamp
                each_ws.cell(row=last_row, column=3).value = ip_address
                each_ws.cell(row=last_row, column=4).value = school_id

                for i, response in enumerate(responses, start=5):
                    each_ws.cell(row=last_row, column=i).value = response

                each_wb.save(excel_file_path2)
                each_wb.close()

                print(f"no: {num+1}, role: { role}, time: {timestamp}, id: {school_id}, ip: {ip_address} 저장이 완료되었습니다.")
                
            except Exception as e:
                print(f"엑셀 처리 중 오류 발생 (개개인별 응답 저장): {e}")

            finally:
                response_queue.task_done()  # 작업 완료 처리

# 백그라운드 스레드 시작
threading.Thread(target=process_queue, daemon=True).start()

def handle_survey_response(request):
    if request.method == 'POST':
        role = request.POST.get('role')
        school_id = request.POST.get('school_id')
        ip_address = get_client_ip(request)

        question_count = {'student': 22, 'parents': 23, 'teacher': 25}[role]
        responses = [
            int(request.POST.get(f'question{i}', 0)) for i in range(1, question_count + 1)
        ]

        # 큐에 응답 추가
        response_queue.put((role, responses, school_id, ip_address))
        print(f"{role, ip_address, school_id, responses} 응답이 큐에 추가되었습니다.")

    return redirect('survey_complete')


########################################학교별 평균###################################################################
## 학생
# 각 학교별 결과 통계 - 학생용 설문조사 (초/중/고 결과 통계) : 문항별 / 영역별 / 전체 통계
def school_student_statistics(request):
    # 지역 데이터 가져오기
    regions = School.objects.values_list('district', flat=True).distinct().order_by('district')
    
    # school_admin일 경우 해당 학교 정보를 세션에서 가져오기
    if request.session.get('role') == 'school_admin':
        school_name = request.session.get('school_name')
        school_level = request.session.get('school_level')
        print(f"school_admin으로 로그인한 학교: {school_name}, 학교급: {school_level}")  # 디버깅

    else:
        # POST 요청 처리
        if request.method == 'POST':
            school_name = request.POST.get('school_name')
            school_level = request.POST.get('school_level')
            print(f"전달된 학교 이름: {school_name}")  # 디버깅: school_name이 제대로 전달되었는지 확인
            print(f"전달된 학교급: {school_level}")  # 디버깅: school_level 값 확인

    excel_file_path = "\\eduWeb\\surveySite\\main\\surveydata\\survey_result_student.xlsx"
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    ws = wb.active

    row_to_static = None
    responses = []
    people_count = None
        
    # 학교명이 있는 열을 찾아서 해당 행을 가져옴
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):  # 2번째 열에 학교명이 있음
        if str(row[0].value).strip() == str(school_name).strip():
            row_to_static= row[0].row
            print(f"학교명 {school_name}가 {row_to_static}번째 행에 있습니다.")  # 디버깅
            break
        
    if row_to_static:
        ### 1. 각 항목별 평균 구하기 - 초중고 같이 해도 됨 (나중에 초등학교 일 경우만 responses 리스트에서 마지막 값 빼버리는거 구현하기)
        # 저장되어 있는 응답 값을 리스트에 불러오기 
        for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=5, max_col=26):
            for value in cell:
                responses.append(value.value or 0)
                
        # 초등학교인 경우 responses 리스트의 마지막 값을 제거
        if school_level == "초등학교":
            responses.pop()  # 마지막 요소 제거
        
        # 응답 인원 불러오기
        for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=27, max_col=27):
            for value in cell:
                people_count = value.value
            
        print(f"{school_name}에 대한 응답값 합산들 : {responses}") #디버깅
        print(f"{school_name}에 대한 응답인원 : {people_count}") #디버깅
        
        # people_count가 0이거나 None인 경우 처리
        if people_count and people_count > 0:
            average_response = [response / people_count for response in responses]  # 각 항목당 평균 구해둔 리스트
        else:
            average_response = [0 for response in responses]  # people_count가 0이면 모든 평균 값을 0으로 설정
            
        print(f"{school_name}에 대한 평균 응답값들 : {average_response}")  # 디버깅
            
            
        ### 2. 영역별 평균 구하기
        question = 1
        section_response = [0, 0, 0]
        
        if school_level == "초등학교":
            # 초등학교 영역별 합산 (1~8, 9~14, 15~21)
            for response in average_response:
                if question <= 8:
                    section_response[0] += response
                elif question <= 14:
                    section_response[1] += response
                else:
                    section_response[2] += response
                    
                question+=1
        
            average_section_response = [
                    round(section_response[0] / 8, 1),
                    round(section_response[1] / 6, 1),
                    round(section_response[2] / 7, 1)
                ]
        
        else:
            # 중/고등학교 기준 영역별 합산 (1~8, 9~15, 16~22)
            for response in average_response:
                if question <= 8:
                    section_response[0] += response
                elif question <= 15:
                    section_response[1] += response
                else:
                    section_response[2] += response
                question += 1

            average_section_response = [
                round(section_response[0] / 8, 1),
                round(section_response[1] / 7, 1),
                round(section_response[2] / 7, 1)
            ]
            
        print(f"{school_name}에 대한 영역별 합산 : {section_response}")  # 디버깅
        print(f"{school_name}에 대한 영역별 평균 : {average_section_response}")  # 디버깅    
        
            
        ### 3. 전체 평균 구하기
        average_total_response = [round(sum(average_section_response) / 3, 1)]
        
        print(f"{school_name}에 대한 전체 평균 : {average_total_response}")  # 디버깅
        
    context = {
        'school_name': school_name,
        'responses': responses,
        'average_response': average_response,
        'average_section_response': average_section_response,
        'average_total_response': average_total_response,
        'people_count': people_count,
        'regions': regions,
    }
    
    # 템플릿을 렌더링하여 HTML로 반환합니다.
    html = render_to_string('statistics_admin_content.html', context)
    
    return JsonResponse({
        'html': html,
        'average_response': average_response,
        'average_section_response': average_section_response,
        'people_count': people_count,
        'average_total_response': average_total_response
    })


## 학부모
# 각 학교별 결과 통계 - 학부모 설문조사 (초/중/고 결과 통계) : 문항별 / 영역별 / 전체 통계
def school_parents_statistics(request):
    # 지역 데이터 가져오기
    regions = School.objects.values_list('district', flat=True).distinct().order_by('district')
    
    # school_admin일 경우 해당 학교 정보를 세션에서 가져오기
    if request.session.get('role') == 'school_admin':
        school_name = request.session.get('school_name')
        school_level = request.session.get('school_level')
        print(f"school_admin으로 로그인한 학교: {school_name}, 학교급: {school_level}")  # 디버깅

    else:
        # POST 요청 처리
        if request.method == 'POST':
            school_name = request.POST.get('school_name')
            school_level = request.POST.get('school_level')
            print(f"전달된 학교 이름: {school_name}")  # 디버깅: school_name이 제대로 전달되었는지 확인
            print(f"전달된 학교급: {school_level}")  # 디버깅: school_level 값 확인

    excel_file_path = "\\eduWeb\\surveySite\\main\\surveydata\\survey_result_parents.xlsx"
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    ws = wb.active

    row_to_static = None
    responses = []
    people_count = None
    
    # 학교명이 있는 열을 찾아서 해당 행을 가져옴
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):  # 2번째 열에 학교명이 있음
        if str(row[0].value).strip() == str(school_name).strip():
            row_to_static= row[0].row
            print(f"학교명 {school_name}가 {row_to_static}번째 행에 있습니다.")  # 디버깅
            break
        
    if row_to_static:
        ### 1. 각 항목별 평균 구하기 - 초중고 같이 해도 됨 (나중에 초등학교 일 경우만 responses 리스트에서 마지막 값 빼버리는거 구현하기)
        # 저장되어 있는 응답 값을 리스트에 불러오기 
        for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=5, max_col=27):
            for value in cell:
                responses.append(value.value or 0)
                
        # 초등학교인 경우 responses 리스트의 마지막 값을 제거
        if school_level == "초등학교":
            responses.pop()  # 마지막 요소 제거
        
        # 응답 인원 불러오기
        for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=28, max_col=28):
            for value in cell:
                people_count = value.value
            
        print(f"{school_name}에 대한 응답값 합산들 : {responses}") #디버깅
        print(f"{school_name}에 대한 응답인원 : {people_count}") #디버깅
        
        # people_count가 0이거나 None인 경우 처리
        if people_count and people_count > 0:
            average_response = [response / people_count for response in responses]  # 각 항목당 평균 구해둔 리스트
        else:
            average_response = [0 for response in responses]  # people_count가 0이면 모든 평균 값을 0으로 설정
            
        print(f"{school_name}에 대한 평균 응답값들 : {average_response}")  # 디버깅
        
        
        ### 2. 영역별 평균 구하기
        question = 1
        section_response = [0, 0, 0]
        
        # (초중고유 설문지 영역 같음)영역별 합산 (1~10, 11~16, 17~23)
        for response in average_response:
            if question <= 10:
                section_response[0] += response
            elif question <= 16:
                section_response[1] += response
            else:
                section_response[2] += response
                
            question+=1
    
        average_section_response = [
                round(section_response[0] / 10, 1),
                round(section_response[1] / 6, 1),
                round(section_response[2] / 7, 1)
            ]
            
        print(f"{school_name}에 대한 영역별 합산 : {section_response}")  # 디버깅
        print(f"{school_name}에 대한 영역별 평균 : {average_section_response}")  # 디버깅    
        
        
        ### 3. 전체 평균 구하기
        average_total_response = [round(sum(average_section_response) / 3, 1)]
        
        print(f"{school_name}에 대한 전체 평균 : {average_total_response}")  # 디버깅
        
    context = {
        'school_name': school_name,
        'responses': responses,
        'average_response': average_response,
        'average_section_response': average_section_response,
        'average_total_response': average_total_response,
        'people_count': people_count,
        'regions': regions,
    }
    
    # 템플릿을 렌더링하여 HTML로 반환합니다.
    html = render_to_string('statistics_admin_content.html', context)
    
    return JsonResponse({
        'html': html,
        'average_response': average_response,
        'average_section_response': average_section_response,
        'people_count': people_count,
        'average_total_response': average_total_response
    })



## 교원
# 각 학교별 결과 통계 - 학부모 설문조사 (초/중/고 결과 통계) : 문항별 / 영역별 / 전체 통계
def school_teacher_statistics(request):
    # 지역 데이터 가져오기
    regions = School.objects.values_list('district', flat=True).distinct().order_by('district')
    
    # school_admin일 경우 해당 학교 정보를 세션에서 가져오기
    if request.session.get('role') == 'school_admin':
        school_name = request.session.get('school_name')
        school_level = request.session.get('school_level')
        print(f"school_admin으로 로그인한 학교: {school_name}, 학교급: {school_level}")  # 디버깅

    else:
        # POST 요청 처리
        if request.method == 'POST':
            school_name = request.POST.get('school_name')
            school_level = request.POST.get('school_level')
            print(f"전달된 학교 이름: {school_name}")  # 디버깅: school_name이 제대로 전달되었는지 확인
            print(f"전달된 학교급: {school_level}")  # 디버깅: school_level 값 확인

    excel_file_path = "\\eduWeb\\surveySite\\main\\surveydata\\survey_result_teacher.xlsx"
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    ws = wb.active

    row_to_static = None
    responses = []
    people_count = None
    
    # 학교명이 있는 열을 찾아서 해당 행을 가져옴
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):  # 2번째 열에 학교명이 있음
        if str(row[0].value).strip() == str(school_name).strip():
            row_to_static= row[0].row
            print(f"학교명 {school_name}가 {row_to_static}번째 행에 있습니다.")  # 디버깅
            break
        
    if row_to_static:
        ### 1. 각 항목별 평균 구하기 - 초중고 같이 해도 됨 (나중에 초등학교 일 경우만 responses 리스트에서 마지막 값 빼버리는거 구현하기)
        # 저장되어 있는 응답 값을 리스트에 불러오기 
        for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=5, max_col=29):
            for value in cell:
                responses.append(value.value or 0)
                
        # 초등학교인 경우 responses 리스트의 마지막 값을 제거
        if school_level == "초등학교":
            responses.pop()  # 마지막 요소 제거
        
        # 응답 인원 불러오기
        for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=30, max_col=30):
            for value in cell:
                people_count = value.value
            
        print(f"{school_name}에 대한 응답값 합산들 : {responses}") #디버깅
        print(f"{school_name}에 대한 응답인원 : {people_count}") #디버깅
        
        # people_count가 0이거나 None인 경우 처리
        if people_count and people_count > 0:
            average_response = [response / people_count for response in responses]  # 각 항목당 평균 구해둔 리스트
        else:
            average_response = [0 for response in responses]  # people_count가 0이면 모든 평균 값을 0으로 설정
            
        print(f"{school_name}에 대한 평균 응답값들 : {average_response}")  # 디버깅
        
        
        ### 2. 영역별 평균 구하기
        question = 1
        section_response = [0, 0, 0]
        
        # (초중고유 설문지 영역 같음)영역별 합산 (1~10, 11~17, 18~25)
        for response in average_response:
            if question <= 10:
                section_response[0] += response
            elif question <= 17:
                section_response[1] += response
            else:
                section_response[2] += response
                
            question+=1
    
        average_section_response = [
                round(section_response[0] / 10, 1),
                round(section_response[1] / 7, 1),
                round(section_response[2] / 8, 1)
            ]
            
        print(f"{school_name}에 대한 영역별 합산 : {section_response}")  # 디버깅
        print(f"{school_name}에 대한 영역별 평균 : {average_section_response}")  # 디버깅    
        
        
        ### 3. 전체 평균 구하기
        average_total_response = [round(sum(average_section_response) / 3, 1)]
        
        print(f"{school_name}에 대한 전체 평균 : {average_total_response}")  # 디버깅
        
    context = {
        'school_name': school_name,
        'responses': responses,
        'average_response': average_response,
        'average_section_response': average_section_response,
        'average_total_response': average_total_response,
        'people_count': people_count,
        'regions': regions,
    }
    
    # 템플릿을 렌더링하여 HTML로 반환합니다.
    html = render_to_string('statistics_admin_content.html', context)
    
    return JsonResponse({
        'html': html,
        'average_response': average_response,
        'average_section_response': average_section_response,
        'people_count': people_count,
        'average_total_response': average_total_response
    })



########################################지역별 평균###################################################################

# 지역별 평균 - 학생용 설문조사 (초/중/고 결과 통계) : 문항별 / 영역별 / 전체 통계
def region_student_statistics(request):        
    # 모든 지역을 가져오되, 지역청 관리자는 본인의 지역만 보게 설정
    if request.session.get('role') == 'regional_admin':
        # 지역청 관리자의 지역을 세션에서 가져옴
        region = request.session.get('region')
        regions = [region]  # 지역청 관리자는 본인의 지역만 선택 가능
    else:
        # 본청 관리자는 모든 지역을 볼 수 있음
        regions = School.objects.values_list('district', flat=True).distinct().order_by('district')

    if request.method == 'POST':
        region = request.POST.get('region')
        school_level = request.POST.get('school_level')
        
        print(f"전달된 지역: {region}")  # 디버깅: 지역 확인
        print(f"전달된 학교급: {school_level}")  # 디버깅: 학교급 확인

        # 선택된 지역과 학교급에 맞는 학교들 조회
        schools_in_region = School.objects.filter(district=region, school_level=school_level)
        
        total_responses = []
        total_people_count = 0
    
        for school in schools_in_region:
            # 각 학교별로 데이터를 엑셀에서 가져오는 방식
            excel_file_path = "\\eduWeb\\surveySite\\main\\surveydata\\survey_result_student.xlsx"
            wb = openpyxl.load_workbook(excel_file_path, data_only=True)
            ws = wb.active

            row_to_static = None
            responses = []
            people_count = None
            
            # 학교명이 있는 열을 찾아서 해당 행을 가져옴
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
                if str(row[0].value).strip() == str(school.school_name).strip():
                    row_to_static = row[0].row
                    break
            
            if row_to_static:
                # 응답 값과 응답 인원 합산
                for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=5, max_col=26):
                    for value in cell:
                        responses.append(value.value or 0)
                
                # 응답 인원 불러오기
                for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=27, max_col=27):
                    for value in cell:
                        people_count = value.value
                
                # 지역 전체 응답 합산
                total_responses.append(responses)
                total_people_count += people_count
                
            # 지역 전체 응답 합산 후 average_response 계산
            if total_people_count > 0 and total_responses:  # 응답이 있는 경우에만 계산
                combined_responses = [sum(x) for x in zip(*total_responses)]  # 문항별 응답 합산
                average_response = [response / total_people_count for response in combined_responses]  # 평균 구하기
            else:
                combined_responses = []  # 응답이 없을 경우 빈 리스트 할당
                average_response = []  # 응답이 없을 경우 빈 리스트 할당
    
        
        # 영역별 통계 계산 (초등, 중등, 고등의 경우를 나눠서)
        section_response = [0, 0, 0]
        question = 1
        
        if school_level == '초등학교':
            for response in average_response:
                if question <= 8:
                    section_response[0] += response
                elif question <= 14:
                    section_response[1] += response
                else:
                    section_response[2] += response
                question += 1
            average_section_response = [round(section_response[0] / 8, 1), round(section_response[1] / 6, 1), round(section_response[2] / 7, 1)]
        
        else:
            for response in average_response:
                if question <= 8:
                    section_response[0] += response
                elif question <= 15:
                    section_response[1] += response
                else:
                    section_response[2] += response
                question += 1
            average_section_response = [round(section_response[0] / 8, 1), round(section_response[1] / 7, 1), round(section_response[2] / 7, 1)]
        
        average_total_response = [round(sum(average_section_response) / 3, 1)]
        
        context = {
            'region': region,
            'school_level': school_level,
            'average_response': average_response,
            'average_section_response': average_section_response,
            'average_total_response': average_total_response,
            'people_count': total_people_count,
            'regions': regions,
        }
        
        html = render_to_string('statistics_admin_content.html', context)
        
        return JsonResponse({
            'html': html,
            'average_response': average_response,
            'average_section_response': average_section_response,
            'people_count': total_people_count,
            'average_total_response': average_total_response
        })
    
    return render(request, 'statistics_admin_region_page.html', {'regions': regions, 'role': request.session.get('role')})


## 학부모
# 지역별 평균 - 학부모 설문조사 (초/중/고 결과 통계) : 문항별 / 영역별 / 전체 통계
def region_parents_statistics(request):
    # 모든 지역을 가져오되, 지역청 관리자는 본인의 지역만 보게 설정
    if request.session.get('role') == 'regional_admin':
        # 지역청 관리자의 지역을 세션에서 가져옴
        region = request.session.get('region')
        regions = [region]  # 지역청 관리자는 본인의 지역만 선택 가능
    else:
        # 본청 관리자는 모든 지역을 볼 수 있음
        regions = School.objects.values_list('district', flat=True).distinct().order_by('district')

    if request.method == 'POST':
        region = request.POST.get('region')
        school_level = request.POST.get('school_level')
        
        print(f"전달된 지역: {region}")  # 디버깅: 지역 확인
        print(f"전달된 학교급: {school_level}")  # 디버깅: 학교급 확인

        # 선택된 지역과 학교급에 맞는 학교들 조회
        schools_in_region = School.objects.filter(district=region, school_level=school_level)
        
        total_responses = []
        total_people_count = 0
    
        for school in schools_in_region:
            # 각 학교별로 데이터를 엑셀에서 가져오는 방식
            excel_file_path = "\\eduWeb\\surveySite\\main\\surveydata\\survey_result_parents.xlsx"
            wb = openpyxl.load_workbook(excel_file_path, data_only=True)
            ws = wb.active

            row_to_static = None
            responses = []
            people_count = None
            
            # 학교명이 있는 열을 찾아서 해당 행을 가져옴
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
                if str(row[0].value).strip() == str(school.school_name).strip():
                    row_to_static = row[0].row
                    break
            
            if row_to_static:
                # 응답 값과 응답 인원 합산
                for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=5, max_col=27):
                    for value in cell:
                        responses.append(value.value or 0)
                
                # 응답 인원 불러오기
                for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=28, max_col=28):
                    for value in cell:
                        people_count = value.value
                
                # 지역 전체 응답 합산
                total_responses.append(responses)
                total_people_count += people_count
            
            ### 1. 항목별 평균 구하기
            # 지역 전체 응답 합산 후 average_response 계산
            if total_people_count > 0 and total_responses:  # 응답이 있는 경우에만 계산
                combined_responses = [sum(x) for x in zip(*total_responses)]  # 문항별 응답 합산
                average_response = [response / total_people_count for response in combined_responses]  # 평균 구하기
            else:
                combined_responses = []  # 응답이 없을 경우 빈 리스트 할당
                average_response = []  # 응답이 없을 경우 빈 리스트 할당
    
        
        ### 2. 영역별 평균 구하기
        # 영역별 통계 계산 (초등, 중등, 고등의 경우를 나눠서)
        section_response = [0, 0, 0]
        question = 1
        
        # (초중고유 설문지 영역 같음)영역별 합산 (1~10, 11~16, 17~23)
        for response in average_response:
            if question <= 10:
                section_response[0] += response
            elif question <= 16:
                section_response[1] += response
            else:
                section_response[2] += response
            question += 1
        average_section_response = [
            round(section_response[0] / 10, 1), 
            round(section_response[1] / 6, 1), 
            round(section_response[2] / 7, 1)
        ]
        
        ### 3. 전체 평균 구하기
        average_total_response = [round(sum(average_section_response) / 3, 1)]
        
        context = {
            'region': region,
            'school_level': school_level,
            'average_response': average_response,
            'average_section_response': average_section_response,
            'average_total_response': average_total_response,
            'people_count': total_people_count,
            'regions': regions,
        }
        
        html = render_to_string('statistics_admin_content.html', context)
        
        return JsonResponse({
            'html': html,
            'average_response': average_response,
            'average_section_response': average_section_response,
            'people_count': total_people_count,
            'average_total_response': average_total_response
        })
    
    return render(request, 'statistics_admin_region_page.html', {'regions': regions, 'role': request.session.get('role')})


## 교원
# 지역별 평균 - 교직원 설문조사 (초/중/고 결과 통계) : 문항별 / 영역별 / 전체 통계
def region_teacher_statistics(request):
    # 모든 지역을 가져오되, 지역청 관리자는 본인의 지역만 보게 설정
    if request.session.get('role') == 'regional_admin':
        # 지역청 관리자의 지역을 세션에서 가져옴
        region = request.session.get('region')
        regions = [region]  # 지역청 관리자는 본인의 지역만 선택 가능
    else:
        # 본청 관리자는 모든 지역을 볼 수 있음
        regions = School.objects.values_list('district', flat=True).distinct().order_by('district')

    if request.method == 'POST':
        region = request.POST.get('region')
        school_level = request.POST.get('school_level')
        
        print(f"전달된 지역: {region}")  # 디버깅: 지역 확인
        print(f"전달된 학교급: {school_level}")  # 디버깅: 학교급 확인

        # 선택된 지역과 학교급에 맞는 학교들 조회
        schools_in_region = School.objects.filter(district=region, school_level=school_level)
        
        total_responses = []
        total_people_count = 0
    
        for school in schools_in_region:
            # 각 학교별로 데이터를 엑셀에서 가져오는 방식
            excel_file_path = "\\eduWeb\\surveySite\\main\\surveydata\\survey_result_teacher.xlsx"
            wb = openpyxl.load_workbook(excel_file_path, data_only=True)
            ws = wb.active

            row_to_static = None
            responses = []
            people_count = None
            
            # 학교명이 있는 열을 찾아서 해당 행을 가져옴
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
                if str(row[0].value).strip() == str(school.school_name).strip():
                    row_to_static = row[0].row
                    break
            
            if row_to_static:
                # 응답 값과 응답 인원 합산
                for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=5, max_col=29):
                    for value in cell:
                        responses.append(value.value or 0)
                
                # 응답 인원 불러오기
                for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=30, max_col=30):
                    for value in cell:
                        people_count = value.value
                
                # 지역 전체 응답 합산
                total_responses.append(responses)
                total_people_count += people_count
            
            ### 1. 항목별 평균 구하기
            # 지역 전체 응답 합산 후 average_response 계산
            if total_people_count > 0 and total_responses:  # 응답이 있는 경우에만 계산
                combined_responses = [sum(x) for x in zip(*total_responses)]  # 문항별 응답 합산
                average_response = [response / total_people_count for response in combined_responses]  # 평균 구하기
            else:
                combined_responses = []  # 응답이 없을 경우 빈 리스트 할당
                average_response = []  # 응답이 없을 경우 빈 리스트 할당
    
        
        ### 2. 영역별 평균 구하기
        section_response = [0, 0, 0]
        question = 1
        
        # (초중고유 설문지 영역 같음)영역별 합산 (1~10, 11~17, 18~25)
        for response in average_response:
            if question <= 10:
                section_response[0] += response
            elif question <= 16:
                section_response[1] += response
            else:
                section_response[2] += response
            question += 1
        average_section_response = [
            round(section_response[0] / 10, 1), 
            round(section_response[1] / 7, 1), 
            round(section_response[2] / 8, 1)
        ]
        
        ### 3. 전체 평균 구하기
        average_total_response = [round(sum(average_section_response) / 3, 1)]
        
        context = {
            'region': region,
            'school_level': school_level,
            'average_response': average_response,
            'average_section_response': average_section_response,
            'average_total_response': average_total_response,
            'people_count': total_people_count,
            'regions': regions,
        }
        
        html = render_to_string('statistics_admin_content.html', context)
        
        return JsonResponse({
            'html': html,
            'average_response': average_response,
            'average_section_response': average_section_response,
            'people_count': total_people_count,
            'average_total_response': average_total_response
        })
    
    return render(request, 'statistics_admin_region_page.html', {'regions': regions, 'role': request.session.get('role')})
    

########################################학교급별 평균###################################################################

## 학생
# 전체 통계 - 학생용 설문조사 (초/중/고 결과 통계) : 문항별 / 영역별 / 전체 통계
def total_student_statistics(request):
    school_levels = ['초등학교', '중학교', '고등학교']  # 각 학교급 정의

    if request.method == 'POST':
        school_level = request.POST.get('school_level')  # 선택된 학교급
        print(f"전달된 학교급: {school_level}")  # 디버깅

        # 선택된 학교급에 맞는 학교들 조회 - SQL에서 실행됨
        schools_in_level = School.objects.filter(school_level=school_level).order_by(Lower('school_name'))
        
        total_responses = []  # 응답 합산을 위한 초기화
        total_people_count = 0
    
        for school in schools_in_level:
            print(f"탐색하고 있는 학교: {school}")  # 디버깅
            # 각 학교별로 데이터를 엑셀에서 가져오는 방식
            excel_file_path = "\\eduWeb\\surveySite\\main\\surveydata\\survey_result_student.xlsx"
            wb = openpyxl.load_workbook(excel_file_path, data_only=True)
            ws = wb.active

            row_to_static = None
            responses = []
            people_count = None
            
            
            # 학교명이 있는 열을 찾아서 해당 행을 가져옴
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
                if str(row[0].value).strip() == str(school.school_name).strip():
                    row_to_static = row[0].row
                    break
            
            if row_to_static:
                # 응답 값
                for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=5, max_col=26):
                    for value in cell:
                        responses.append(value.value or 0)
                        
                print(f"응답값: {responses}")  # 디버깅
                
                # 응답 인원 불러오기
                for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=27, max_col=27):
                    for value in cell:
                        people_count = value.value
                        
                print(f"응답인원: {people_count}")  # 디버깅
                
                # 지역 전체 응답 합산 (total_responses에 각 school's 응답을 더함)
                if not total_responses:
                    total_responses = responses  # 첫 번째 학교의 응답값으로 초기화
                else:
                    total_responses = [x + y for x, y in zip(total_responses, responses)]  # 같은 인덱스끼리 더함
                
                # 전체 응답자 수 더함
                total_people_count += people_count
                
                print(f"응답값합산: {total_responses}")  # 디버깅
                print(f"사람수 합산: {total_people_count}")  # 디버깅
                
        
        # 학교 전체 응답 합산 후 average_response 계산
        if total_people_count > 0 and total_responses:  # 응답이 있는 경우에만 계산
            average_response = [response / total_people_count for response in total_responses]  # 평균 구하기
        else:
            combined_responses = []
            average_response = []
        
        # 영역별 통계 계산
        section_response = [0, 0, 0]
        question = 1

        if school_level == '초등학교':
            for response in average_response:
                if question <= 8:
                    section_response[0] += response
                elif question <= 14:
                    section_response[1] += response
                else:
                    section_response[2] += response
                question += 1
            average_section_response = [round(section_response[0] / 8, 1), round(section_response[1] / 6, 1), round(section_response[2] / 7, 1)]

        else:  # 중학교, 고등학교
            for response in average_response:
                if question <= 8:
                    section_response[0] += response
                elif question <= 15:
                    section_response[1] += response
                else:
                    section_response[2] += response
                question += 1
            average_section_response = [round(section_response[0] / 8, 1), round(section_response[1] / 7, 1), round(section_response[2] / 7, 1)]

        average_total_response = [round(sum(average_section_response) / 3, 1)]
        
        context = {
            'school_level': school_level,
            'average_response': average_response,
            'average_section_response': average_section_response,
            'people_count': total_people_count,
            'average_total_response': average_total_response,
        }
        
        html = render_to_string('statistics_total_content.html', context)
        
        return JsonResponse({
            'html': html,
            'average_response': average_response,
            'average_section_response': average_section_response,
            'people_count': total_people_count,
            'average_total_response': average_total_response
        })
    
    return render(request, 'statistics_admin_total_page.html', {'school_levels': school_levels})


## 학부모
# 전체 통계 - 학부모용 설문조사 (초/중/고 결과 통계) : 문항별 / 영역별 / 전체 통계
def total_parents_statistics(request):
    school_levels = ['초등학교', '중학교', '고등학교']  # 각 학교급 정의

    if request.method == 'POST':
        school_level = request.POST.get('school_level')  # 선택된 학교급
        print(f"전달된 학교급: {school_level}")  # 디버깅

        # 선택된 학교급에 맞는 학교들 조회 - SQL에서 실행됨
        schools_in_level = School.objects.filter(school_level=school_level).order_by(Lower('school_name'))
        
        total_responses = []  # 응답 합산을 위한 초기화
        total_people_count = 0
    
        for school in schools_in_level:
            print(f"탐색하고 있는 학교: {school}")  # 디버깅
            # 각 학교별로 데이터를 엑셀에서 가져오는 방식
            excel_file_path = "\\eduWeb\\surveySite\\main\\surveydata\\survey_result_parents.xlsx"
            wb = openpyxl.load_workbook(excel_file_path, data_only=True)
            ws = wb.active

            row_to_static = None
            responses = []
            people_count = None
            
            
            # 학교명이 있는 열을 찾아서 해당 행을 가져옴
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
                if str(row[0].value).strip() == str(school.school_name).strip():
                    row_to_static = row[0].row
                    break
            
            if row_to_static:
                # 응답 값
                for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=5, max_col=27):
                    for value in cell:
                        responses.append(value.value or 0)
                        
                print(f"응답값: {responses}")  # 디버깅
                
                # 응답 인원 불러오기
                for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=28, max_col=28):
                    for value in cell:
                        people_count = value.value
                        
                print(f"응답인원: {people_count}")  # 디버깅
                
                # 지역 전체 응답 합산 (total_responses에 각 school 응답을 더함)
                if not total_responses:
                    total_responses = responses  # 첫 번째 학교의 응답값으로 초기화
                else:
                    total_responses = [x + y for x, y in zip(total_responses, responses)]  # 같은 인덱스끼리 더함
                
                # 전체 응답자 수 더함
                total_people_count += people_count
                
                print(f"응답값합산: {total_responses}")  # 디버깅
                print(f"사람수 합산: {total_people_count}")  # 디버깅
                
        ### 1. 항목별 평균 구하기
        # 학교 전체 응답 합산 후 average_response 계산
        if total_people_count > 0 and total_responses:  # 응답이 있는 경우에만 계산
            average_response = [response / total_people_count for response in total_responses]  # 평균 구하기
        else:
            combined_responses = []
            average_response = []
        
        ### 2. 영역별 평균 구하기
        # 영역별 통계 계산
        section_response = [0, 0, 0]
        question = 1

        # (초중고유 설문지 영역 같음)영역별 합산 (1~10, 11~16, 17~23)
        for response in average_response:
            if question <= 10:
                section_response[0] += response
            elif question <= 16:
                section_response[1] += response
            else:
                section_response[2] += response
            question += 1
        average_section_response = [
            round(section_response[0] / 10, 1), 
            round(section_response[1] / 6, 1), 
            round(section_response[2] / 7, 1)
        ]

        ### 3. 전체 평균 구하기
        average_total_response = [round(sum(average_section_response) / 3, 1)]
        
        context = {
            'school_level': school_level,
            'average_response': average_response,
            'average_section_response': average_section_response,
            'people_count': total_people_count,
            'average_total_response': average_total_response,
        }
        
        html = render_to_string('statistics_total_content.html', context)
        
        return JsonResponse({
            'html': html,
            'average_response': average_response,
            'average_section_response': average_section_response,
            'people_count': total_people_count,
            'average_total_response': average_total_response
        })
    
    return render(request, 'statistics_admin_total_page.html', {'school_levels': school_levels})

## 교원
# 전체 통계 - 교직원용 설문조사 (초/중/고 결과 통계) : 문항별 / 영역별 / 전체 통계
def total_teacher_statistics(request):
    school_levels = ['초등학교', '중학교', '고등학교']  # 각 학교급 정의

    if request.method == 'POST':
        school_level = request.POST.get('school_level')  # 선택된 학교급
        print(f"전달된 학교급: {school_level}")  # 디버깅

        # 선택된 학교급에 맞는 학교들 조회 - SQL에서 실행됨
        schools_in_level = School.objects.filter(school_level=school_level).order_by(Lower('school_name'))
        
        total_responses = []  # 응답 합산을 위한 초기화
        total_people_count = 0
    
        for school in schools_in_level:
            print(f"탐색하고 있는 학교: {school}")  # 디버깅
            # 각 학교별로 데이터를 엑셀에서 가져오는 방식
            excel_file_path = "\\eduWeb\\surveySite\\main\\surveydata\\survey_result_teacher.xlsx"
            wb = openpyxl.load_workbook(excel_file_path, data_only=True)
            ws = wb.active

            row_to_static = None
            responses = []
            people_count = None
            
            
            # 학교명이 있는 열을 찾아서 해당 행을 가져옴
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
                if str(row[0].value).strip() == str(school.school_name).strip():
                    row_to_static = row[0].row
                    break
            
            if row_to_static:
                # 응답 값
                for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=5, max_col=29):
                    for value in cell:
                        responses.append(value.value or 0)
                        
                print(f"응답값: {responses}")  # 디버깅
                
                # 응답 인원 불러오기
                for cell in ws.iter_cols(min_row=row_to_static, max_row=row_to_static, min_col=30, max_col=30):
                    for value in cell:
                        people_count = value.value
                        
                print(f"응답인원: {people_count}")  # 디버깅
                
                # 지역 전체 응답 합산 (total_responses에 각 school 응답을 더함)
                if not total_responses:
                    total_responses = responses  # 첫 번째 학교의 응답값으로 초기화
                else:
                    total_responses = [x + y for x, y in zip(total_responses, responses)]  # 같은 인덱스끼리 더함
                
                # 전체 응답자 수 더함
                total_people_count += people_count
                
                print(f"응답값합산: {total_responses}")  # 디버깅
                print(f"사람수 합산: {total_people_count}")  # 디버깅
                
        ### 1. 항목별 평균 구하기
        # 학교 전체 응답 합산 후 average_response 계산
        if total_people_count > 0 and total_responses:  # 응답이 있는 경우에만 계산
            average_response = [response / total_people_count for response in total_responses]  # 평균 구하기
        else:
            combined_responses = []
            average_response = []
        
        ### 2. 영역별 평균 구하기
        # 영역별 통계 계산
        section_response = [0, 0, 0]
        question = 1

        # (초중고유 설문지 영역 같음)영역별 합산 (1~10, 11~17, 18~25)
        for response in average_response:
            if question <= 10:
                section_response[0] += response
            elif question <= 16:
                section_response[1] += response
            else:
                section_response[2] += response
            question += 1
        average_section_response = [
            round(section_response[0] / 10, 1), 
            round(section_response[1] / 7, 1), 
            round(section_response[2] / 8, 1)
        ]
        
        ### 3. 전체 평균 구하기
        average_total_response = [round(sum(average_section_response) / 3, 1)]
        
        context = {
            'school_level': school_level,
            'average_response': average_response,
            'average_section_response': average_section_response,
            'people_count': total_people_count,
            'average_total_response': average_total_response,
        }
        
        html = render_to_string('statistics_total_content.html', context)
        
        return JsonResponse({
            'html': html,
            'average_response': average_response,
            'average_section_response': average_section_response,
            'people_count': total_people_count,
            'average_total_response': average_total_response
        })
    
    return render(request, 'statistics_admin_total_page.html', {'school_levels': school_levels})
    


########################################### 설문 질문 리스트 ####################################################
#초등학생용 설문
def ele_stuSur_question(request):
    role = request.GET.get('role', 'student')
    
    # 학교문화 관련 질문 (1~8)
    options = ["매우 아니다", "아니다", "보통이다", "그렇다", "매우 그렇다"]
    school_culture = [
        {'id': 1, 'text': ''},
        {'id': 2, 'text': ''},
        {'id': 3, 'text': ''},
        {'id': 4, 'text': ''},
        {'id': 5, 'text': ''},
        {'id': 6, 'text': ''},
        {'id': 7, 'text': ''},
        {'id': 8, 'text': ''}
    ]

    # 학교구조 관련 질문 (9~14)
    school_structure = [
        {'id': 9, 'text': ''},
        {'id': 10, 'text': ''},
        {'id': 11, 'text': ''},
        {'id': 12, 'text': ''},
        {'id': 13, 'text': ''},
        {'id': 14, 'text': ''}
    ]

    # 민주시민교육 실천 관련 질문 (15~21)
    democratic_citizenship = [
        {'id': 15, 'text': ''},
        {'id': 16, 'text': ''},
        {'id': 17, 'text': ''},
        {'id': 18, 'text': ''},
        {'id': 19, 'text': ''},
        {'id': 20, 'text': ''},
        {'id': 21, 'text': ''}
    ]

    return render(request, 'survey_ele_student.html', {
        'school_culture': school_culture,
        'school_structure': school_structure,
        'democratic_citizenship': democratic_citizenship,
        'options': options,
        'school_id': request.session.get('school_id'),
        'role': role
    })


# 중고등학생용 설문
def midHigh_stuSur_question(request):
    role = request.GET.get('role', 'student')
    
    # 학교문화 관련 질문 (1~8)
    school_culture = [
        {'id': 1, 'text': ''},
        {'id': 2, 'text': ''},
        {'id': 3, 'text': ''},
        {'id': 4, 'text': ''},
        {'id': 5, 'text': ''},
        {'id': 6, 'text': ''},
        {'id': 7, 'text': ''},
        {'id': 8, 'text': ''}
    ]

    # 학교구조 관련 질문 (9~15)
    school_structure = [
        {'id': 9, 'text': ''},
        {'id': 10, 'text': ''},
        {'id': 11, 'text': ''},
        {'id': 12, 'text': ''},
        {'id': 13, 'text': ''},
        {'id': 14, 'text': ''},
        {'id': 15, 'text': ''}
    ]

    # 민주시민교육 실천 관련 질문 (16~22)
    democratic_citizenship = [
        {'id': 16, 'text': ''},
        {'id': 17, 'text': ''},
        {'id': 18, 'text': ''},
        {'id': 19, 'text': ''},
        {'id': 20, 'text': ''},
        {'id': 21, 'text': ''},
        {'id': 22, 'text': ''}
    ]

    options = ["매우 아니다", "아니다", "보통이다", "그렇다", "매우 그렇다"]

    return render(request, 'survey_midHigh_student.html', {
        'school_culture': school_culture,
        'school_structure': school_structure,
        'democratic_citizenship': democratic_citizenship,
        'options': options,
        'school_id': request.session.get('school_id'),
        'role': role
    })



# 유치원 학부모용 설문
def kinder_parSur_question(request):
    role = request.GET.get('role', 'parents')
    # 학교문화 관련 질문 (1~10)
    school_culture = [
        {'id': 1, 'text': ''},
        {'id': 2, 'text': ''},
        {'id': 3, 'text': ''},
        {'id': 4, 'text': ''},
        {'id': 5, 'text': ''},
        {'id': 6, 'text': ''},
        {'id': 7, 'text': ''},
        {'id': 8, 'text': ''},
        {'id': 9, 'text': ''},
        {'id': 10, 'text': ''}
    ]

    # 학교구조 관련 질문 (11~16)
    school_structure = [
        {'id': 11, 'text': ''},
        {'id': 12, 'text': ''},
        {'id': 13, 'text': ''},
        {'id': 14, 'text': ''},
        {'id': 15, 'text': ''},
        {'id': 16, 'text': ''}
    ]

    # 민주시민교육 실천 관련 질문 (17~23)
    democratic_citizenship = [
        {'id': 17, 'text': ''},
        {'id': 18, 'text': ''},
        {'id': 19, 'text': ''},
        {'id': 20, 'text': ''},
        {'id': 21, 'text': ''},
        {'id': 22, 'text': ''},
        {'id': 23, 'text': ''}
    ]

    options = ["매우 아니다", "아니다", "보통이다", "그렇다", "매우 그렇다"]

    return render(request, 'survey_kinder_parents.html', {
        'school_culture': school_culture,
        'school_structure': school_structure,
        'democratic_citizenship': democratic_citizenship,
        'options': options,
        'school_id' : request.session.get('school_id'),
        'role': role
    })



# 초중고 학부모용 설문
def school_parSur_question(request):
    role = request.GET.get('role', 'parents')
    
    # 학교문화 관련 질문 (1~10)
    school_culture = [
        {'id': 1, 'text': ''},
        {'id': 2, 'text': ''},
        {'id': 3, 'text': ''},
        {'id': 4, 'text': ''},
        {'id': 5, 'text': ''},
        {'id': 6, 'text': ''},
        {'id': 7, 'text': ''},
        {'id': 8, 'text': ''},
        {'id': 9, 'text': ''},
        {'id': 10, 'text': ''}
    ]

    # 학교구조 관련 질문 (11~16)
    school_structure = [
        {'id': 11, 'text': ''},
        {'id': 12, 'text': ''},
        {'id': 13, 'text': ''},
        {'id': 14, 'text': ''},
        {'id': 15, 'text': ''},
        {'id': 16, 'text': ''}
    ]

    # 민주시민교육 실천 관련 질문 (17~23)
    democratic_citizenship = [
        {'id': 17, 'text': ''},
        {'id': 18, 'text': ''},
        {'id': 19, 'text': ''},
        {'id': 20, 'text': ''},
        {'id': 21, 'text': ''},
        {'id': 22, 'text': ''},
        {'id': 23, 'text': ''}
    ]

    options = ["매우 아니다", "아니다", "보통이다", "그렇다", "매우 그렇다"]

    return render(request, 'survey_school_parents.html', {
        'school_culture': school_culture,
        'school_structure': school_structure,
        'democratic_citizenship': democratic_citizenship,
        'options': options,
        'school_id' : request.session.get('school_id'),
        'role': role
    })



# 유치원 교직원용 설문
def kinder_teaSur_question(request):
    role = request.GET.get('role', 'teacher')
    
    # 학교문화 관련 질문 (1~10)
    school_culture = [
        {'id': 1, 'text': ''},
        {'id': 2, 'text': ''},
        {'id': 3, 'text': ''},
        {'id': 4, 'text': ''},
        {'id': 5, 'text': ''},
        {'id': 7, 'text': ''},
        {'id': 8, 'text': ''},
        {'id': 9, 'text': ''},
        {'id': 10, 'text': ''}
    ]

    # 학교구조 관련 질문 (11~17)
    school_structure = [
        {'id': 11, 'text': ''},
        {'id': 12, 'text': ''},
        {'id': 13, 'text': ''},
        {'id': 14, 'text': ''},
        {'id': 15, 'text': ''},
        {'id': 16, 'text': ''},
        {'id': 17, 'text': ''}
    ]

    # 민주시민교육 실천 관련 질문 (18~25)
    democratic_citizenship = [
        {'id': 18, 'text': ''},
        {'id': 19, 'text': ''},
        {'id': 20, 'text': ''},
        {'id': 21, 'text': ''},
        {'id': 22, 'text': ''},
        {'id': 23, 'text': ''},
        {'id': 24, 'text': ''},
        {'id': 25, 'text': ''}
    ]

    options = ["매우 아니다", "아니다", "보통이다", "그렇다", "매우 그렇다"]

    return render(request, 'survey_kinder_teacher.html', {
        'school_culture': school_culture,
        'school_structure': school_structure,
        'democratic_citizenship': democratic_citizenship,
        'options': options,
        'school_id' : request.session.get('school_id'),
        'role': role
    })



# 초중고 교직원용 설문
def school_teaSur_question(request):
    role = request.GET.get('role', 'teacher')
    
    # 학교문화 관련 질문 (1~10)
    school_culture = [
        {'id': 1, 'text': ''},
        {'id': 2, 'text': ''},
        {'id': 3, 'text': ''},
        {'id': 4, 'text': ''},
        {'id': 5, 'text': ''},
        {'id': 6, 'text': ''},
        {'id': 7, 'text': ''},
        {'id': 8, 'text': ''},
        {'id': 9, 'text': ''},
        {'id': 10, 'text': ''}
    ]

    # 학교구조 관련 질문 (11~17)
    school_structure = [
        {'id': 11, 'text': ''},
        {'id': 12, 'text': ''},
        {'id': 13, 'text': ''},
        {'id': 14, 'text': ''},
        {'id': 15, 'text': ''},
        {'id': 16, 'text': ''},
        {'id': 17, 'text': ''}
    ]

    # 민주시민교육 실천 관련 질문 (18~25)
    democratic_citizenship = [
        {'id': 18, 'text': ''},
        {'id': 19, 'text': ''},
        {'id': 20, 'text': ''},
        {'id': 21, 'text': ''},
        {'id': 22, 'text': ''},
        {'id': 23, 'text': ''},
        {'id': 24, 'text': ''},
        {'id': 25, 'text': ''}
    ]

    options = ["매우 아니다", "아니다", "보통이다", "그렇다", "매우 그렇다"]

    return render(request, 'survey_school_teacher.html', {
        'school_culture': school_culture,
        'school_structure': school_structure,
        'democratic_citizenship': democratic_citizenship,
        'options': options,
        'school_id': request.session.get('school_id'),
        'role': role
    })