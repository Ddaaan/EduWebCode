# from django.core.management.base import BaseCommand
# import openpyxl
# from main.models import School
# from django.contrib.auth.hashers import make_password  # 비밀번호 해시화를 위한 함수

# class Command(BaseCommand):
#     help = '학교 아이디를 기준으로 학교 이름을 엑셀 파일을 통해 업데이트합니다.'

#     def add_arguments(self, parser):
#         parser.add_argument('excel_file', type=str, help="D:\Daeun\eduWeb\surveySite\main\surveydata\data.xlsx")

#     def handle(self, *args, **kwargs):
#         excel_file = kwargs['excel_file']
#         wb = openpyxl.load_workbook(excel_file)
#         ws = wb.active

#         for row in ws.iter_rows(min_row=2, values_only=True):
#             school_id = row[7]  # 아이디가 8번째 열에 있다고 가정
#             new_school_name = row[3]  # 새 학교 이름이 4번째 열에 있다고 가정
            
#             try:
#                 school = School.objects.get(school_id=school_id)
#                 school.school_name = new_school_name  # 학교 이름 업데이트
#                 school.save()
#                 self.stdout.write(self.style.SUCCESS(f'ID {school_id}: 학교 이름이 "{new_school_name}"(으)로 성공적으로 변경되었습니다.'))
#             except School.DoesNotExist:
#                 self.stdout.write(self.style.ERROR(f'ID {school_id}: 해당하는 학교를 찾을 수 없습니다.'))

from django.core.management.base import BaseCommand
import openpyxl
from main.models import School
from django.contrib.auth.hashers import make_password  # 비밀번호 해시화를 위한 함수

class Command(BaseCommand):
    help = '학교 아이디를 기준으로 비밀번호를 엑셀 파일을 통해 업데이트합니다.'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help="D:\Daeun\eduWeb\surveySite\main\surveydata\data.xlsx")

    def handle(self, *args, **kwargs):
        excel_file = kwargs['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            school_id = row[7]  # 아이디가 8번째 열에 있다고 가정
            new_password = row[8]  # 새 비밀번호가 9번째 열에 있다고 가정
            
            try:
                school = School.objects.get(school_id=school_id)
                if new_password:
                    school.school_pw = make_password(new_password)  # 비밀번호 해쉬화 후 업데이트
                school.save()
                self.stdout.write(self.style.SUCCESS(f'ID {school_id}: 비밀번호가 성공적으로 변경되었습니다.'))
            except School.DoesNotExist:
                self.stdout.write(self.style.ERROR(f'ID {school_id}: 해당하는 학교를 찾을 수 없습니다.'))